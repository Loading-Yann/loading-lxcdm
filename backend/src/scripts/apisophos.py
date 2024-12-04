import requests
import pandas as pd
import json
import os
import openpyxl
from datetime import datetime

# Fonction pour lire les credentials depuis credentials.json
def read_credentials():
    with open('credentials.json', 'r') as file:
        return json.load(file)

# Fonction pour lire ou initialiser le fichier firewalls.json
def read_or_initialize_firewalls():
    if os.path.exists('firewalls.json'):
        with open('firewalls.json', 'r') as file:
            return json.load(file)
    else:
        return []

# Fonction pour mettre à jour firewalls.json sans doublons
def update_firewalls(firewalls, client_name, date_created, firewall_data):
    existing_entry = next((entry for entry in firewalls if entry["clientName"] == client_name), None)

    if existing_entry:
        existing_entry["firewalls"] = firewall_data
        existing_entry["dateCreated"] = date_created
    else:
        firewalls.append({
            "clientName": client_name,
            "dateCreated": date_created,
            "firewalls": firewall_data
        })

    with open('firewalls.json', 'w') as file:
        json.dump(firewalls, file, indent=4)

# Fonction pour récupérer les données des pare-feu via l'API Sophos
def fetch_firewall_data(client_id, client_secret):
    base_url = "https://api.central.sophos.com"
    token_url = "https://id.sophos.com/api/v2/oauth2/token"
    whoami_url = base_url + "/whoami/v1"

    auth_data = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'token'
    }

    try:
        response = requests.post(token_url, data=auth_data)
        response.raise_for_status()
        access_token = response.json().get("access_token")

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }

        response = requests.get(whoami_url, headers=headers)
        response.raise_for_status()
        tenant_id = response.json()["id"]
        data_region = response.json()["apiHosts"]["dataRegion"]

        firewall_url = f"{data_region}/firewall/v1/firewalls"
        headers["X-Tenant-ID"] = tenant_id

        response = requests.get(firewall_url, headers=headers)
        response.raise_for_status()
        return response.json()

    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except Exception as err:
        print(f"Other error occurred: {err}")

    return None

# Fonction pour modifier une feuille Excel
def modify_excel_sheet(sheet, client_name, data_length):
    # Ajouter la colonne A avec le nom du client
    sheet.insert_cols(1)
    sheet.cell(row=1, column=1, value="client_name")
    for i in range(2, data_length + 2):
        sheet.cell(row=i, column=1, value=client_name)

    # Ajouter la colonne "version"
    sheet.insert_cols(9)
    sheet.cell(row=1, column=9, value="version")
    for i in range(2, data_length + 2):
        sheet.cell(row=i, column=9, value=f'=RIGHT(H{i}, LEN(H{i}) - SEARCH("#", SUBSTITUTE(H{i}, "_", "#", LEN(H{i}) - LEN(SUBSTITUTE(H{i}, "_", "")))))')

    # Ajouter la colonne "modèle"
    sheet.insert_cols(11)
    sheet.cell(row=1, column=11, value="modèle")
    for i in range(2, data_length + 2):
        sheet.cell(row=i, column=11, value=f'=LEFT(J{i}, SEARCH("_", J{i}) - 1)')

    # Ajuster la largeur des colonnes
    for col in sheet.columns:
        col_letter = col[0].column_letter
        current_width = sheet.column_dimensions[col_letter].width or 8.43
        sheet.column_dimensions[col_letter].width = current_width * 2

    # Masquer les colonnes inutiles
    for col_letter in ['B', 'D', 'E', 'H', 'J', 'L', 'M', 'O']:
        if col_letter in sheet.column_dimensions:
            sheet.column_dimensions[col_letter].hidden = True

# Fonction pour créer le fichier Excel pour chaque client
def create_excel_file(firewall_data, client_name, folder_name):
    df_client = pd.json_normalize(firewall_data)
    excel_filename = os.path.join(folder_name, f"firewalls_sophos_{client_name}.xlsx")
    df_client.to_excel(excel_filename, index=False)

    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook.active

    modify_excel_sheet(sheet, client_name, len(firewall_data))

    workbook.save(excel_filename)
    workbook.close()

    print(f"Fichier Excel créé pour {client_name} : {excel_filename}")

# Fonction pour créer le fichier combiné
def create_combined_excel(folder_name):
    combined_data = []

    # Parcourir tous les fichiers Excel du dossier
    for filename in os.listdir(folder_name):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(folder_name, filename)
            df = pd.read_excel(filepath)
            combined_data.append(df)

    # Combiner les données et sauvegarder dans un fichier Excel
    combined_df = pd.concat(combined_data, ignore_index=True)
    combined_filename = os.path.join(folder_name, "firewalls_sophos_combined.xlsx")
    combined_df.to_excel(combined_filename, index=False)

    # Modifier le fichier combiné avec OpenPyXL
    workbook = openpyxl.load_workbook(combined_filename)
    sheet = workbook.active

    # Appeler la fonction pour modifier la feuille Excel
    modify_excel_sheet(sheet, "combined", len(combined_df))

    # Supprimer la colonne A
    sheet.delete_cols(1)

    # Masquer les colonnes C et K
    sheet.column_dimensions['C'].hidden = True
    sheet.column_dimensions['K'].hidden = True

    # Modifier la formule de la colonne H (en fonction de la colonne I)
    for i in range(2, len(combined_df) + 2):
        sheet.cell(row=i, column=8, value=f'=RIGHT(I{i}, LEN(I{i}) - SEARCH("#", SUBSTITUTE(I{i}, "_", "#", LEN(I{i}) - LEN(SUBSTITUTE(I{i}, "_", "")))))')

    # Afficher la colonne H
    sheet.column_dimensions['H'].hidden = False

    # Modifier la formule de la colonne J (en fonction de la colonne N)
    for i in range(2, len(combined_df) + 2):
        sheet.cell(row=i, column=10, value=f'=LEFT(N{i}, SEARCH("_", N{i}) - 1)')

    # Afficher la colonne J
    sheet.column_dimensions['J'].hidden = False

    # Tri de l'entière feuille en fonction de la colonne A (qui est maintenant la première colonne)
    data = list(sheet.iter_rows(min_row=2, values_only=True))
    data.sort(key=lambda row: row[0])  # Trier en fonction de la valeur de la colonne A (index 0)

    # Réécrire les lignes triées dans la feuille
    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)

    # Sauvegarder et fermer le fichier
    workbook.save(combined_filename)
    workbook.close()

    print(f"Fichier Excel combiné créé, modifié et trié : {combined_filename}")

# Fonction principale
def main():
    credentials = read_credentials()
    firewalls = read_or_initialize_firewalls()
    folder_name = f"sophos_excel_{datetime.now().strftime('%Y-%m-%d')}"

    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    for cred in credentials:
        client_id = cred["clientId"]
        client_secret = cred["clientSecret"]
        client_name = cred["clientName"]
        date_created = cred["dateCreated"]

        firewall_data = fetch_firewall_data(client_id, client_secret)

        if firewall_data:
            create_excel_file(firewall_data["items"], client_name, folder_name)
            update_firewalls(firewalls, client_name, date_created, firewall_data["items"])

    create_combined_excel(folder_name)

if __name__ == "__main__":
    main()
